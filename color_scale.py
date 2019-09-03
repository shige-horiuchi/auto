import random
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule

wb = Workbook() # ワークブックを作成
ws = wb.active  # ワークシートを取得

values = random.sample(range(100), k = 20) # 100までの数字で、20個をランダムに取得
for i, value in enumerate(values, 1):
    column_a = ws.cell(i, 1)
    column_a.value = value   # A列に値をセット
    column_b = ws.cell(i, 2)
    column_b.value = value   # B列に値をセット

row_length = len(values)

# 2色スケールの設定
two_color_scale = ColorScaleRule(
    start_type = 'min', start_color = 'FFFFCC',
    end_type = 'max', end_color = 'FF8000'
)
ws.conditional_formatting.add(
    f'A1:A{row_length}', two_color_scale
)

# 3色スケールの設定
three_color_scale = ColorScaleRule(
    start_type = 'percentile', start_value = 10, start_color = 'FFFFCC',
    mid_type = 'percentile', mid_value = 50, mid_color = 'FF8000',
    end_type = 'percentile', end_value = 90, end_color = 'FF0000'
)

# B列に3色スケールを設定
ws.conditional_formatting.add(
    f'B1:B{row_length}', three_color_scale
)

ws.title = 'カラースケール' # ワークシート名
wb.save('color_scale.xlsx') # Excelファイル保存