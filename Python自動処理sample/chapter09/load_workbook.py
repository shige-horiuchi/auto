from openpyxl import load_workbook

filename = 'various_worksheets.xlsx'
wb = load_workbook(filename, read_only=True)        #←ワークブックを読み込む
print(f'{filename} のワークシート情報を読み込みます')

for ws in wb.worksheets:        #←すべてのワークシートの列と行の情報を調べる
    print(f'ワークシート名: {ws.title}')
    print(f'- 列の値 最小:{ws.min_column}, 最大:{ws.max_column}')
    print(f'- 行の値 最小:{ws.min_row}, 最大:{ws.max_row}')
