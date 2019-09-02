from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

wb = Workbook()   # ワークブックを作成
ws = wb.active    # 選択されているワークシートを取得
navy_colored_font = Font(name = 'ＭＳ　Ｐ明朝', size = 20, color = '000080') # フォント、サイズ、色を指定

a1 = ws['A1']
a1.value = 'フォントを変える'
a1.font = navy_colored_font # セルのフォント設定

salmon_colored_fill = PatternFill('solid', fgColor = 'FA8072') # セルの塗りつぶし設定

a2 = ws['A2']
a2.value = '背景色を変える'
a2.fill = salmon_colored_fill # セルの塗りつぶし設定

red_colored_thin = Side(style = 'thin', color = 'FF0000')
blue_colored_double = Side(style = 'double', color = '0000FF')
border = Border(
    left = red_colored_thin, right = red_colored_thin,
    top = blue_colored_double, bottom = blue_colored_double
)

b4 = ws['B4']              # B4のセルを指定
b4.value = '罫線で囲む'     # B4のセルに「罫線で囲む」を書き込む
b4.border = border         # セルの罫線

ws.title = 'セル書き込みワークシート' # ワークシートの名前を設定
wb.save('write_cell_with.xlsx')          # ワークブック（Excelファイル）を保存