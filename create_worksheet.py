from openpyxl import Workbook

wb = Workbook()
ws0 = wb.active
ws0.title = 'ワークシート0'
ws1 = wb.create_sheet('ワークシート1')
ws2 = wb.create_sheet('ワークシート2')
ws3 = wb.create_sheet('sheet1')
print(f'ワークシート： {wb.sheetnames}')

wb.active = 3
wb.save('new_worksheets.xlsx')