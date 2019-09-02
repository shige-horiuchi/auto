from datetime import datetime
from openpyxl import Workbook

wb = Workbook()   # ワークブックを作成
ws = wb.active    # 選択されているワークシートを取得

ws['A1'] = 'セル1' # A1のセルに直接書き込む

b1 = ws['B1']      # B1のセルを指定
b1.value = 3.1     # B1のセルに「3.1」を書き込む

c1 = ws['C1']             # C1のセルを指定
c1.value = datetime.now() # C1のセルに「日付時刻」を書き込む

ws.title = 'セル書き込みワークシート' # ワークシートの名前を設定
wb.save('write_cell.xlsx')          # ワークブック（Excelファイル）を保存